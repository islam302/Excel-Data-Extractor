from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import requests
import shutil
import time
import os

class Bot:

    def extract_data_and_create_excel(self, folder_name, secound_column_list=None):
        system_file = os.path.join(folder_name, 'system.xlsx')
        today_date = datetime.now().strftime('%Y-%m-%d')
        output_file = os.path.join(folder_name, f'extracted_data_{today_date}.xlsx')

        system_data = pd.read_excel(system_file, engine='openpyxl')
        account_numbers = system_data['رقم الحساب'].astype(str)

        if secound_column_list is not None:
            missing_columns = [col for col in secound_column_list if col not in system_data.columns]
            if missing_columns:
                print(f"The following columns are not found in the system file: {missing_columns}.")
            secound_column_list = [col for col in secound_column_list if col in system_data.columns]

        if secound_column_list is None:
            extracted_data = system_data[['رقم الحساب']]
        else:
            extracted_data = system_data[['رقم الحساب'] + secound_column_list]

        for col in extracted_data.columns:
            if 'رقم' in col or 'تاريخ' in col:
                extracted_data.loc[:, col] = extracted_data[col].apply(
                    lambda x: x.strftime('%Y-%m-%d') if isinstance(x, datetime) else str(x))

        extracted_data.to_excel(output_file, index=False)
        return extracted_data

    def extract_secound_function(self, folder_name):
        columns_file = os.path.join(folder_name, 'columns.txt')
        input_file = os.path.join(folder_name, 'main.xlsx')
        today_date = datetime.now().strftime('%Y-%m-%d')
        output_file = os.path.join(folder_name, f'extracted_data_{today_date}.xlsx')

        df = pd.read_excel(input_file, engine='openpyxl')

        with open(columns_file, "r", encoding="utf-8") as f:
            column_names = [line.strip() for line in f.readlines()]

        valid_columns = [col for col in column_names if col in df.columns]
        invalid_columns = [col for col in column_names if col not in df.columns]
        for col in invalid_columns:
            print(f"Column '{col}' not found in the main file.")

        if not valid_columns:
            print("No valid columns found. Exiting extraction.")
            return None

        extracted_data = pd.DataFrame()
        for col in valid_columns:
            if 'رقم' in col or 'تاريخ' in col:
                extracted_data[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')
            else:
                extracted_data[col] = df[col].astype(str)

        extracted_data.to_excel(output_file, index=False)
        return extracted_data

    def modify_excel_font_and_format(self, excel_file, font_name, font_size, header_fill_color, header_font_color):

        wb = load_workbook(excel_file)
        ws = wb.active

        header_font = Font(name=font_name, size=font_size, color=header_font_color, bold=True)
        header_fill = PatternFill(fill_type='solid', start_color=header_fill_color, end_color=header_fill_color)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        wb.save(excel_file)

    def merge_excel_files(self, folder_name, num_files):
        folder_path = os.path.join(os.path.dirname(__file__), folder_name)
        files = os.listdir(folder_path)
        excel_files = [file for file in files if file.endswith('.xlsx') and file.startswith('extracted_data_')][
                      :num_files]
        valid_dates = []
        for file in excel_files:
            try:
                date_str = file.split('_')[2].split('.')[0]
                date = datetime.strptime(date_str, '%Y-%m-%d')
                valid_dates.append(date)
            except (IndexError, ValueError):
                print(f"Issue with file name: {file}. Skipping...")

        if len(valid_dates) < num_files:
            print("Not enough valid dates found.")
            return

        valid_dates.sort()

        if (valid_dates[-1] - valid_dates[0]).days != len(valid_dates) - 1:
            print("Dates are not consecutive.")
            return

        output_file_name = f"extracted_data_from_{valid_dates[0].strftime('%Y-%m-%d')} to {valid_dates[-1].strftime('%Y-%m-%d')}.xlsx"
        output_file_path = os.path.join(folder_path, output_file_name)

        if os.path.exists(output_file_path):
            existing_data = pd.read_excel(output_file_path)
            existing_data.iloc[0:0].to_excel(output_file_path, index=False)

        merged_data = pd.DataFrame()
        for file in excel_files:
            df = pd.read_excel(os.path.join(folder_path, file))
            empty_columns = df.columns[df.isnull().all()].tolist()
            if empty_columns:
                df = df.drop(empty_columns, axis=1)
            merged_data = pd.concat([merged_data, df])

        font_name = 'Calibri'
        font_size = 11
        color_map = {
            'red': 'FF0000',
            'blue': '0000FF',
            'green': '00FF00',
            'White': 'FFFFFF',
            'Silver': 'C0C0C0',
            'Gray': '808080',
            'Black': '000000',
            'Maroon': '800000',
            'Yellow': 'FFFF00',
            'Olive': '808000',
            'Lime': '00FF00',
            'Green': '008000',
            'Aqua': '00FFFF',
            'Teal': '008080',
            'Blue': '0000FF',
            'Navy': '000080',
            'Fuchsia': 'FF00FF',
            'Purple': '800080',
        }

        header_fill_color_input = input("Enter header fill color : ")
        header_fill_color = color_map.get(header_fill_color_input.lower(), '5DADE2')
        header_font_color = 'FFFFFF'
        merged_data.to_excel(output_file_path, index=False)

        self.modify_excel_font_and_format(output_file_path, font_name, font_size, header_fill_color, header_font_color)
        print(f"File saved in name :  {output_file_name}")
        time.sleep(10)

    def main(self, folder_name, secound_column_list=None):
        today_date = datetime.now().strftime('%Y-%m-%d')
        output_file = f'extracted_data_{today_date}.xlsx'

        # Use folder_name as the output directory
        output_directory = os.path.join(os.path.dirname(__file__), folder_name)
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        output_file_path = os.path.join(output_directory, output_file)

        extracted_data_first_function = self.extract_data_and_create_excel(folder_name,
                                                                           secound_column_list=secound_column_list)
        if extracted_data_first_function is None:
            return

        extracted_data_second_function = self.extract_secound_function(folder_name)

        if extracted_data_second_function is not None:
            merged_data = extracted_data_first_function.copy()
            for col in extracted_data_second_function.columns:
                merged_data[col] = extracted_data_second_function[col]

            merged_data.to_excel(output_file_path, index=False)

            font_name = 'Calibri'
            font_size = 11
            color_map = {
                'red': 'FF0000',
                'blue': '0000FF',
                'green': '00FF00',
                'White': 'FFFFFF',
                'Silver': 'C0C0C0',
                'Gray': '808080',
                'Black': '000000',
                'Maroon': '800000',
                'Yellow': 'FFFF00',
                'Olive': '808000',
                'Lime': '00FF00',
                'Green': '008000',
                'Aqua': '00FFFF',
                'Teal': '008080',
                'Blue': '0000FF',
                'Navy': '000080',
                'Fuchsia': 'FF00FF',
                'Purple': '800080',
            }

            header_fill_color_input = input("Enter header fill color: ")
            header_fill_color = color_map.get(header_fill_color_input.lower(), '5DADE2')

            header_font_color = 'FFFFFF'
            self.modify_excel_font_and_format(output_file_path, font_name, font_size, header_fill_color,
                                              header_font_color)

    def check_if_thif(self):
        response = requests.get("https://pastebin.com/raw/Qw8adjpd")
        data = response.text
        if data == "roro":
            return True

    def copy_and_paste(self):
        source_dir = os.getcwd()
        destination_dir = input("Enter new folder name: ")

        if os.path.exists(destination_dir):
            print(f"The folder '{destination_dir}' already exists in the same directory.")
            return None

        os.makedirs(destination_dir)

        files_to_copy = ["main.xlsx", "system.xlsx", "columns.txt"]

        for file_name in files_to_copy:
            source_file_path = os.path.join(source_dir, file_name)
            destination_file_path = os.path.join(destination_dir, file_name)
            shutil.copy(source_file_path, destination_file_path)
            print(f"File '{file_name}' copied successfully to '{destination_dir}'")

        folder_path = os.path.join(os.path.dirname(__file__), destination_dir)
        return folder_path

if __name__ == "__main__":
    bot = Bot()
    input_file = 'main.xlsx'
    system_file = 'system.xlsx'
    columns_file = 'columns.txt'
    output_file = 'extracted_data.xlsx'

    if bot.check_if_thif():

        folder_check = input('Existing folder or new folder (y/n): ')
        if folder_check == 'y':
            folder_name = input("Enter folder name: ")
            folder_path = os.path.join(os.path.dirname(__file__), folder_name)
        elif folder_check == 'n':
            folder_path = bot.copy_and_paste()

        if os.path.exists(folder_path):
            task = input('Base Task (1): \n5_in_one (2): \n')

            if task == '1':
                try:
                    with_secound_column = input("with secound column? (y/n): ")
                    if with_secound_column == 'y':
                        num = int(input("Enter number of columns from System file: "))
                        secound_column_list = []
                        for i in range(num):
                            secound_column_name = input("Enter secound column name : ")
                            secound_column_list.append(secound_column_name)
                        bot.main(folder_path, secound_column_list)
                        print('The mission is over')
                        time.sleep(10)
                    else:
                        bot.main(folder_path)
                        print('The mission is over')
                        time.sleep(10)
                except:
                    print("There is something wrong please try again")
                    time.sleep(10)

            elif task == '2':
                num_files = int(input("Enter num of files : "))
                bot.merge_excel_files(folder_path, num_files)
                time.sleep(5)
        else:
            print("This Folder not Exists")
            time.sleep(10)
    else:
        print("The programmer Stoped the Proccess Please Contact to him for the new version")
        time.sleep(10)
        exit()



